import os
import json
import csv
import yaml
from typing import Optional
from pypdf import PdfReader
import docx
import openpyxl
from app.core.logging import logger

class UniversalParser:
    """
    Universal Parser to extract raw text content from various file formats.
    """
    
    def extract_text(self, file_path: str) -> str:
        """
        Detect file extension and route to appropriate extraction method.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found at path: {file_path}")
            
        _, ext = os.path.splitext(file_path.lower())
        
        logger.info(f"Extracting text from: {file_path} (format: {ext})")
        
        if ext == ".pdf":
            return self._parse_pdf(file_path)
        elif ext == ".docx":
            return self._parse_docx(file_path)
        elif ext in (".xlsx", ".xls"):
            return self._parse_excel(file_path)
        elif ext in (".md", ".markdown"):
            return self._parse_text_file(file_path)
        elif ext in (".txt", ".text"):
            return self._parse_text_file(file_path)
        elif ext == ".csv":
            return self._parse_csv(file_path)
        elif ext in (".json", ".yaml", ".yml"):
            return self._parse_structured_file(file_path)
        else:
            # Fallback to plain text read
            try:
                return self._parse_text_file(file_path)
            except Exception as e:
                raise ValueError(f"Unsupported file format '{ext}' and failed to read as plain text: {e}")

    def _parse_pdf(self, file_path: str) -> str:
        reader = PdfReader(file_path)
        text_parts = []
        for i, page in enumerate(reader.pages):
            content = page.extract_text()
            if content:
                text_parts.append(content)
        return "\n".join(text_parts)

    def _parse_docx(self, file_path: str) -> str:
        doc = docx.Document(file_path)
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        return "\n".join(text_parts)

    def _parse_excel(self, file_path: str) -> str:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(values_only=True):
                if any(val is not None for val in row):
                    row_str = ", ".join([str(val) if val is not None else "" for val in row])
                    text_parts.append(row_str)
        return "\n".join(text_parts)

    def _parse_text_file(self, file_path: str) -> str:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    def _parse_csv(self, file_path: str) -> str:
        text_parts = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                if row:
                    text_parts.append(", ".join(row))
        return "\n".join(text_parts)

    def _parse_structured_file(self, file_path: str) -> str:
        # For JSON/YAML, load and dump as formatted string for LLM readability
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
            
        _, ext = os.path.splitext(file_path.lower())
        if ext == ".json":
            try:
                data = json.loads(content)
                return json.dumps(data, indent=2)
            except Exception:
                return content
        else:
            try:
                data = yaml.safe_load(content)
                return yaml.dump(data, default_flow_style=False)
            except Exception:
                return content

universal_parser = UniversalParser()
